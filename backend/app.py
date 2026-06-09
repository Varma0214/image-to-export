from flask import Flask, request, jsonify
from flask_cors import CORS
import pytesseract
from PIL import Image, ImageEnhance, ImageFilter, ImageOps
import io
import base64
import re
import platform
import numpy as np
from collections import Counter

# PDF
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch

# Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}})

if platform.system() == "Windows":
    pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"


# ─────────────────────────────────────────────
# IMAGE PREPROCESSING
# ─────────────────────────────────────────────

def preprocess_image(image):
    """Multi-stage enhancement for best OCR accuracy on tables."""
    if image.mode == 'RGBA':
        # White background composite
        bg = Image.new('RGB', image.size, (255, 255, 255))
        bg.paste(image, mask=image.split()[3])
        image = bg
    elif image.mode != 'RGB':
        image = image.convert('RGB')

    # Upscale small images — Tesseract needs ≥300 DPI equivalent
    w, h = image.size
    if w < 1600:
        scale = 1600 / w
        new_w, new_h = int(w * scale), int(h * scale)
        image = image.resize((new_w, new_h), Image.LANCZOS)

    # Enhance contrast strongly
    image = ImageEnhance.Contrast(image).enhance(2.5)
    # Sharpen edges
    image = ImageEnhance.Sharpness(image).enhance(3.0)
    # Convert to grayscale
    image = image.convert('L')
    # Auto-level (stretch histogram)
    image = ImageOps.autocontrast(image, cutoff=2)
    # Apply unsharp mask for crisper characters
    image = image.filter(ImageFilter.UnsharpMask(radius=1, percent=150, threshold=3))

    return image


# ─────────────────────────────────────────────
# COLUMN / ROW CLUSTERING  (KEY FIX)
# ─────────────────────────────────────────────

def cluster_1d(values, gap_threshold):
    """Cluster sorted values into groups where consecutive gap ≤ threshold."""
    if not values:
        return []
    sorted_vals = sorted(set(values))
    clusters = [[sorted_vals[0]]]
    for v in sorted_vals[1:]:
        if v - clusters[-1][-1] <= gap_threshold:
            clusters[-1].append(v)
        else:
            clusters.append([v])
    return clusters


def detect_columns_from_whitespace(words, img_width):
    """
    Detect column boundaries by finding vertical white-space gaps across
    ALL rows, not just by clustering word starts.
    
    Strategy: Build a horizontal 'word coverage' array.
    Column separators are x-ranges with zero word coverage.
    """
    # Build coverage array at pixel resolution (downsampled by 2)
    scale = 2
    arr_w = img_width // scale + 1
    coverage = np.zeros(arr_w, dtype=np.int32)

    for w in words:
        l = w['left'] // scale
        r = (w['right'] + 1) // scale
        l = max(0, min(l, arr_w - 1))
        r = max(0, min(r, arr_w - 1))
        coverage[l:r] += 1

    # Find gap regions (zero coverage)
    in_gap = coverage == 0
    # Require gap to be at least 'min_gap_px' pixels wide (in original coords)
    min_gap_px = 15  # pixels in original image
    min_gap_scaled = max(2, min_gap_px // scale)

    # Find contiguous gap segments
    gap_starts = []
    gap_ends = []
    i = 0
    while i < arr_w:
        if in_gap[i]:
            start = i
            while i < arr_w and in_gap[i]:
                i += 1
            end = i - 1
            if (end - start + 1) >= min_gap_scaled:
                gap_starts.append(start * scale)
                gap_ends.append(end * scale)
        else:
            i += 1

    if not gap_starts:
        # No clear gaps — fall back to word-start clustering
        return None, None

    # Column boundaries: before first gap start, between gaps, after last gap end
    col_starts = [0] + [e + 1 for e in gap_ends]
    col_ends = [s - 1 for s in gap_starts] + [img_width]

    return col_starts, col_ends


def assign_word_to_col(word_left, word_right, col_starts, col_ends):
    """Assign a word to its best column based on overlap."""
    word_mid = (word_left + word_right) / 2
    best_ci = 0
    best_overlap = -1
    for ci, (cs, ce) in enumerate(zip(col_starts, col_ends)):
        overlap_start = max(word_left, cs)
        overlap_end = min(word_right, ce)
        overlap = max(0, overlap_end - overlap_start)
        if overlap > best_overlap:
            best_overlap = overlap
            best_ci = ci
    # If no overlap, find nearest column by midpoint
    if best_overlap == 0:
        dists = [abs(word_mid - (cs + ce) / 2) for cs, ce in zip(col_starts, col_ends)]
        best_ci = dists.index(min(dists))
    return best_ci


# ─────────────────────────────────────────────
# MAIN TABLE EXTRACTION
# ─────────────────────────────────────────────

def extract_table_with_bbox(image):
    """
    Robust table extraction using Tesseract bounding boxes.
    Uses whitespace-gap column detection to correctly split
    columns that have a large horizontal gap (e.g. 'Name' | 'class').
    """
    # Run with best LSTM engine, treat as uniform block of text
    configs_to_try = [
        r'--oem 1 --psm 6',   # LSTM, assume uniform block
        r'--oem 1 --psm 4',   # LSTM, assume single column of variable-size text
        r'--oem 0 --psm 6',   # Legacy engine fallback
    ]

    data = None
    for cfg in configs_to_try:
        try:
            d = pytesseract.image_to_data(image, config=cfg, output_type=pytesseract.Output.DICT)
            valid_count = sum(1 for t, c in zip(d['text'], d['conf'])
                              if str(t).strip() and int(str(c).lstrip('-') or -1) > 20)
            if valid_count >= 2:
                data = d
                break
        except Exception:
            continue

    if data is None:
        raw_text = pytesseract.image_to_string(image, config='--psm 6')
        return [], raw_text

    # ── Collect valid words ──
    words = []
    n = len(data['text'])
    for i in range(n):
        text = str(data['text'][i]).strip()
        try:
            conf = int(data['conf'][i])
        except (ValueError, TypeError):
            conf = -1
        if text and conf > 15:
            left = int(data['left'][i])
            top = int(data['top'][i])
            width = max(1, int(data['width'][i]))
            height = max(1, int(data['height'][i]))
            words.append({
                'text': text,
                'left': left,
                'top': top,
                'width': width,
                'height': height,
                'right': left + width,
                'bottom': top + height,
                'conf': conf,
            })

    if not words:
        raw_text = pytesseract.image_to_string(image, config='--psm 6')
        return [], raw_text

    img_width = image.size[0]

    # ── Step 1: Row clustering ──
    avg_height = sum(w['height'] for w in words) / len(words)
    row_gap_threshold = avg_height * 0.55

    tops = sorted(set(w['top'] for w in words))
    row_clusters = cluster_1d(tops, row_gap_threshold)
    top_to_row = {v: ri for ri, cluster in enumerate(row_clusters) for v in cluster}
    num_rows = len(row_clusters)

    rows_words = [[] for _ in range(num_rows)]
    for w in words:
        ri = top_to_row.get(w['top'])
        if ri is not None:
            rows_words[ri].append(w)

    for ri in range(num_rows):
        rows_words[ri].sort(key=lambda w: w['left'])

    # ── Step 2: Column detection via whitespace gaps ──
    col_starts, col_ends = detect_columns_from_whitespace(words, img_width)

    if col_starts is None:
        # Fallback: cluster left-x values, but with a larger gap threshold
        all_lefts = [w['left'] for w in words]
        avg_char_width = sum(w['width'] / max(len(w['text']), 1) for w in words) / len(words)
        # Use a larger threshold — at least 3× average char width
        col_gap = max(avg_char_width * 3, 20)
        left_clusters = cluster_1d(all_lefts, col_gap)
        num_cols_fb = len(left_clusters)
        col_starts = [min(c) for c in left_clusters]
        col_ends = col_starts[1:] + [img_width]

    num_cols = len(col_starts)

    # ── Step 3: Build grid ──
    grid = [[""] * num_cols for _ in range(num_rows)]

    for ri, row_ws in enumerate(rows_words):
        for w in row_ws:
            ci = assign_word_to_col(w['left'], w['right'], col_starts, col_ends)
            if grid[ri][ci]:
                grid[ri][ci] += " " + w['text']
            else:
                grid[ri][ci] = w['text']

    # ── Step 4: Clean grid ──
    # Remove fully-empty rows
    grid = [row for row in grid if any(c.strip() for c in row)]
    if not grid:
        raw_text = pytesseract.image_to_string(image, config='--psm 6')
        return [], raw_text

    # Pad to uniform width
    max_cols = max(len(r) for r in grid)
    grid = [r + [""] * (max_cols - len(r)) for r in grid]

    # Remove columns empty in every row
    non_empty_cols = [ci for ci in range(max_cols)
                      if any(grid[ri][ci].strip() for ri in range(len(grid)))]
    if non_empty_cols:
        grid = [[row[ci] for ci in non_empty_cols] for row in grid]

    raw_text = pytesseract.image_to_string(image, config='--oem 1 --psm 6')
    return grid, raw_text


def fallback_line_parse(text):
    """Parse raw OCR text into a table by splitting on tabs or multi-spaces."""
    lines = [l.strip() for l in text.strip().split('\n') if l.strip()]
    if not lines:
        return []

    table = []
    for line in lines:
        if '\t' in line:
            cols = [c.strip() for c in line.split('\t') if c.strip()]
        else:
            cols = re.split(r'\s{2,}', line)
            cols = [c.strip() for c in cols if c.strip()]
        if cols:
            table.append(cols)

    if not table:
        return []

    col_counts = Counter(len(r) for r in table)
    target_cols = col_counts.most_common(1)[0][0]

    normalized = []
    for row in table:
        if len(row) >= target_cols:
            normalized.append(row[:target_cols])
        else:
            normalized.append(row + [""] * (target_cols - len(row)))
    return normalized


# ─────────────────────────────────────────────
# ROUTES
# ─────────────────────────────────────────────

@app.route('/ocr', methods=['POST'])
def ocr_image():
    req_data = request.get_json()
    if not req_data or 'image' not in req_data:
        return jsonify({'error': 'No image provided'}), 400

    try:
        img_data = req_data['image']
        if ',' in img_data:
            img_data = img_data.split(',')[1]
        img_bytes = base64.b64decode(img_data)
        image = Image.open(io.BytesIO(img_bytes))

        processed = preprocess_image(image.copy())

        table, raw_text = extract_table_with_bbox(processed)

        # If bbox gave only 1 column or nothing, try line-parse fallback
        if not table or max((len(r) for r in table), default=0) <= 1:
            fallback = fallback_line_parse(raw_text)
            if fallback and max(len(r) for r in fallback) > max((len(r) for r in table), default=0):
                table = fallback

        if not table:
            table = [[raw_text.strip() or "No text detected"]]

        # Normalize column count
        max_c = max(len(r) for r in table)
        table = [r + [""] * (max_c - len(r)) for r in table]

        return jsonify({
            'raw_text': raw_text,
            'table': table,
            'rows': len(table),
            'cols': max_c,
        })

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/export/pdf', methods=['POST'])
def export_pdf():
    data = request.get_json()
    table_data = data.get('table', [])
    title = data.get('title', 'Extracted Data')

    if not table_data:
        return jsonify({'error': 'No table data'}), 400

    try:
        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            leftMargin=0.75 * inch, rightMargin=0.75 * inch,
            topMargin=0.75 * inch, bottomMargin=0.75 * inch,
        )
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Heading1'],
            fontSize=16, spaceAfter=14,
            textColor=colors.HexColor('#1a1a2e'),
        )
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.15 * inch))

        max_cols = max(len(row) for row in table_data)
        # Ensure every row is the same length
        norm = [list(row) + [''] * (max_cols - len(row)) for row in table_data]

        # Each cell as a Paragraph so text wraps properly
        cell_style = ParagraphStyle(
            'Cell', parent=styles['Normal'],
            fontSize=9, leading=12,
            wordWrap='CJK',
        )
        header_cell_style = ParagraphStyle(
            'HCell', parent=styles['Normal'],
            fontSize=9, leading=12,
            textColor=colors.white,
            fontName='Helvetica-Bold',
            wordWrap='CJK',
        )

        def make_cell(text, is_header):
            st = header_cell_style if is_header else cell_style
            return Paragraph(str(text) if text else '', st)

        para_norm = []
        for ri, row in enumerate(norm):
            para_norm.append([make_cell(cell, ri == 0) for cell in row])

        page_w = A4[0] - 1.5 * inch
        col_w = page_w / max_cols

        t = Table(para_norm, colWidths=[col_w] * max_cols, repeatRows=1)
        t.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            # Alternating rows
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [colors.white, colors.HexColor('#f0f4ff')]),
            # Alignment & padding
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#c0c8e0')),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#1a1a2e')),
        ]))
        elements.append(t)

        doc.build(elements)
        buf.seek(0)
        pdf_b64 = base64.b64encode(buf.read()).decode('utf-8')
        return jsonify({'pdf': pdf_b64, 'filename': f'{title}.pdf'})

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/export/excel', methods=['POST'])
def export_excel():
    data = request.get_json()
    table_data = data.get('table', [])
    title = data.get('title', 'Exported Data')

    if not table_data:
        return jsonify({'error': 'No table data'}), 400

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]

        header_fill = PatternFill("solid", fgColor="1A1A2E")
        alt_fill = PatternFill("solid", fgColor="EEF2FF")
        header_font = Font(bold=True, color="FFFFFF", size=11, name='Calibri')
        body_font = Font(size=10, name='Calibri')
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(style='thin', color='C0C8E0')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for r_idx, row in enumerate(table_data, start=1):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=str(val) if val else '')
                cell.border = border
                cell.alignment = center
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                else:
                    cell.font = body_font
                    if r_idx % 2 == 0:
                        cell.fill = alt_fill

        # Auto-fit column widths
        for col in ws.columns:
            max_len = 0
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 24

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        xl_b64 = base64.b64encode(buf.read()).decode('utf-8')
        return jsonify({'excel': xl_b64, 'filename': f'{title}.xlsx'})

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


if __name__ == '__main__':
    app.run(debug=True, port=5000)